from openpyxl import load_workbook # import the module openpyxl

print "Connecting to database....."
wb=load_workbook("/home/subham/Desktop/SpamClassifier/UCIDataset/databaseUCI.xlsx")
ws = wb.active
print "Database Connected!!"

col_header = ['A','B','C','D','E','F','G','H','I','J','K','L','M','N','O','P','Q','R','S','T','U','V','W','X','Y','Z','AA','AB','AC','AD','AE','AF','AG','AH','AI','AJ','AK','AL','AM','AN','AO','AP','AQ','AR','AS','AT','AU','AV','AW','AX','AY','AZ','BA','BB','BC','BD','BE','BF',]



# Reading the keywords file
file = open("/home/subham/Desktop/SpamClassifier/UCIDataset/Keywords.txt","r")
keyword_list = file.read().split()
print keyword_list
num_keyword = len(keyword_list)
print num_keyword




column = ws['BF']
spam_status = []
keyword_spam_occur = [0]*num_keyword
keyword_ham_occur = [0]*num_keyword
keyword_ham_prob = [0]*num_keyword
keyword_spam_prob = [0]*num_keyword





ham = 0  # Stores the number of ham(Non- spam) messages in database
spam = 0 #  Stores the number of spam messages in database
total = len(column) # stores the total no. of messages in the database


# Access database and storing the contents of spam_status column i.e. BF in list called spam_status
for x in xrange(len(column)): 
    spam_status.append(column[x].value)
    if(column[x].value == 0):
    	ham = ham+1
    else:
    	spam = spam+1
print ""
print "Total messages =",total
print "No. of spam messages =",spam
print "No. of ham messages =",ham
print ""
# Find the probability that how many times each keyword occurs in a spam / non-spam mail
print "Calculating Keyword Probabilities......"
for i in range(0,num_keyword):
	column = ws[col_header[i]]
	for x in xrange(len(column)):
		if(column[x].value>0 and spam_status[x]==1 ):
			keyword_spam_occur[i] = keyword_spam_occur[i] +1
		elif(column[x].value>0 and spam_status[x]==0 ):
			keyword_ham_occur[i] = keyword_ham_occur[i] +1
	keyword_ham_prob[i] =  keyword_ham_occur[i]/float(ham)
	keyword_spam_prob[i] = keyword_spam_occur[i]/float(spam)


print "Done!!"
print ""
print "Generating the probability Table : "
print ""
print '%-12s%-12s%-12s%-12s%-12s' % ("Keyword", "#SpamOccur", "SpamProb", "#HamOccur", "HamProb")  

for i in range(0,num_keyword):
	print '%-12s%-12i%-12f%-12i%-12f' % (keyword_list[i], keyword_spam_occur[i],keyword_spam_prob[i],keyword_ham_occur[i],keyword_ham_prob[i])

print "\nThe Classifier Algorithm has been successfully trained.....PHASE -I complete"

print "\n Please enter a String to predict spam probabilty & ham probability :"
choice = 0
while(choice==0):
	text = raw_input().split()
	occur_bool = [0]*num_keyword
	for x in range(0,num_keyword):
		if(keyword_list[x] in text):
			occur_bool[x] = 1 
	print occur_bool
	print ""
	test_spam_prob = 1
	test_ham_prob = 1
	for x in range(0,num_keyword):
		if(occur_bool[x]==1):
			test_spam_prob = test_spam_prob * (keyword_spam_prob[x])
			test_ham_prob = test_ham_prob * (keyword_ham_prob[x])
		else:
			test_spam_prob = test_spam_prob * (1-keyword_spam_prob[x])
			test_ham_prob = test_ham_prob * (1-keyword_ham_prob[x])
	total_prob = test_spam_prob* (float(spam)/float(total)) + test_ham_prob* (float(ham)/float(total))
	print "Spam Probabilty : ", test_spam_prob * (float(spam)/float(total)) / total_prob
	print "Ham Probabilty  : ", test_ham_prob * (float(ham)/float(total)) /total_prob

